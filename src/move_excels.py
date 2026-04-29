"""Move Excel files into matching patent folders based on patent IDs."""

import argparse
import shutil
from pathlib import Path

from openpyxl import load_workbook


def read_first_column_values(excel_path: Path) -> set[str]:
    """Read all non-empty values from the first column of an Excel file."""
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active

    values = set()

    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row, column=1).value

        if value is not None and str(value).strip():
            values.add(str(value).strip())

    workbook.close()
    return values


def move_excels_by_patent_id(source_dir: str) -> None:
    """Move Excel files into patent folders when a matching ID is found."""
    source_path = Path(source_dir)

    if not source_path.exists():
        raise FileNotFoundError(f"Directory not found: {source_dir}")

    excel_files = [
        file for file in source_path.glob("*.xls*")
        if file.is_file() and not file.name.startswith("~$")
    ]

    if not excel_files:
        print(f"No Excel files found in {source_dir}")
        return

    moved = 0
    skipped = 0

    for excel_file in excel_files:
        try:
            patent_ids = read_first_column_values(excel_file)

            matching_folders = [
                patent_id for patent_id in patent_ids
                if (source_path / patent_id).is_dir()
            ]

            if not matching_folders:
                print(f"Skipped {excel_file.name}: no matching patent folder found")
                skipped += 1
                continue

            target_folder = source_path / matching_folders[0]
            target_path = target_folder / excel_file.name

            shutil.move(str(excel_file), str(target_path))

            print(f"Moved: {excel_file.name} -> {target_folder.name}/")
            moved += 1

        except Exception as error:
            print(f"Failed to process {excel_file.name}: {error}")
            skipped += 1

    print("\nExcel organization summary")
    print(f"Moved files: {moved}")
    print(f"Skipped files: {skipped}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Move Excel files into matching patent folders."
    )
    parser.add_argument("source_dir", help="Directory containing Excel files and patent folders")
    args = parser.parse_args()

    move_excels_by_patent_id(args.source_dir)


if __name__ == "__main__":
    main()
